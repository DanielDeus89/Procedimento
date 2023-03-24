#
# 
import openpyxl
import os
from openpyxl.utils import get_column_letter
import datetime
import shutil


directory = ''

if os.path.exists("FBR-EN-4-448"):
    shutil.rmtree('FBR-EN-4-448')

if not os.path.exists("FBR-EN-4-448"):
    os.mkdir('FBR-EN-4-448')

print("Estou trabalhando na criação dos arquivos agora. Por favor, aguarde um pouco enquanto eu os crio. Isso pode levar alguns minutos. Obrigado pela sua paciência!")

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".xlsx" e se o nome começa com "SmarX"
    if filename.endswith("xlsx") and filename.startswith("Smar") or filename.startswith("Adda"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # remove o arquivo
        os.remove(current_file)

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".txt"
    if filename.endswith("xlsx"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # constrói o novo caminho completo do arquivo
        new_file = os.path.join(directory, 'FBR-EN-4-448.xlsx')
        # renomeia o arquivo
        os.rename(current_file, new_file)        


meses = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
mes_atual = datetime.date.today().month


FT1_MP1 = ['SmarX1125A2599', 'SmarX1125A3432', 'SmarX1125A4769', 'SmarX1125A5821','SmarX1125A8970']
FT2_MP1 = ['SmarX1188A5001', 'SmarX1188A5002', 'SmarX1188A8006', 'SmarX1188A9179', 'SmarX8700A8116', 'SmarX8700A2018']
FT2_MP2 = ['SmarX1125A1276', 'SmarX1125A2827', 'SmarX1125A2837', 'SmarX1125A7804']
FT2_MP3 = ['SmarX1125A1278', 'SmarX1125A2848', 'SmarX1125A7800', 'SmarX1125A8047']
MGT_MP6 = ['SmarX1125A5825', 'SmarX1125A7192']
ST_MP1 = ['SmarX1280A4702', 'SmarX1280A0375']
BURN_IN = ['AddaX8754A0140','AddaX8754A0145']
ST_MP3 = ['SmarX1288A0067']
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in FT1_MP1:                     
    planilha["C4"] = 'FT1-MP1'
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in FT2_MP1:                     
    planilha["C4"] = 'FT2-MP1' 
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in FT2_MP2:                     
    planilha["C4"] = 'FT2-MP2'
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in MGT_MP6:                     
    planilha["C4"] = 'MGT-MP6' 
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in ST_MP1:                     
    planilha["C4"] = 'ST-MP1'

    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
    
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in FT2_MP3:                     
    planilha["C4"] = 'FT2-MP3'

    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()

arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in BURN_IN:                     
    planilha["C4"] = 'BURN IN'

    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
   
arquivo = openpyxl.load_workbook("FBR-EN-4-448.xlsx")
planilha = arquivo.active

for id in ST_MP3:                     
    planilha["C4"] = 'ST-MP3'

    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()
   

for filename in os.listdir():
    if filename.startswith('FBR-EN-4-448.xlsx'):
        os.remove('FBR-EN-4-448.xlsx')


for filename in os.listdir():
    if filename.endswith('.xlsx'):
        os.rename(filename, os.path.join('FBR-EN-4-448', filename))
       
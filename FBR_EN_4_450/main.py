import openpyxl
import os
from openpyxl.utils import get_column_letter
import datetime
import shutil


directory = ''

if os.path.exists("FBR-EN-4-450"):
    shutil.rmtree('FBR-EN-4-450')

if not os.path.exists("FBR-EN-4-450"):
    os.mkdir('FBR-EN-4-450')

print("Estou trabalhando na criação dos arquivos agora. Por favor, aguarde um pouco enquanto eu os crio. Isso pode levar alguns minutos. Obrigado pela sua paciência!")

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".xlsx" e se o nome começa com "SmarX"
    if filename.endswith("xlsx") and filename.startswith("Smar") or filename.startswith("Adda"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # remove o arquivo
        os.remove(current_file)

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".txt"
    if filename.endswith("xlsx"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # constrói o novo caminho completo do arquivo
        new_file = os.path.join(directory, 'FBR-EN-4-450.xlsx')
        # renomeia o arquivo
        os.rename(current_file, new_file)        


meses = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
mes_atual = datetime.date.today().month

FT1_MP1 = ['SmarX1125A2599', 'SmarX1125A3432', 'SmarX1125A4769', 'SmarX1125A5821','SmarX1125A8970']
FT2_MP2 = ['SmarX1125A1276', 'SmarX1125A2827', 'SmarX1125A2837', 'SmarX1125A7804']
FT2_MP3 = ['SmarX1125A1278', 'SmarX1125A2848', 'SmarX1125A7800', 'SmarX1125A8047']
    
arquivo = openpyxl.load_workbook("FBR-EN-4-450.xlsx")
planilha = arquivo.active

for id in FT1_MP1:                     
    planilha["B4"] = 'FT1-MP1'
    planilha["D4"] = id   
    planilha["G4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close() 

arquivo = openpyxl.load_workbook("FBR-EN-4-450.xlsx")
planilha = arquivo.active

for id in FT2_MP2:                     
    planilha["B4"] = 'FT2-MP2'
    planilha["D4"] = id   
    planilha["G4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close() 

arquivo = openpyxl.load_workbook("FBR-EN-4-450.xlsx")
planilha = arquivo.active

for id in FT2_MP3:                     
    planilha["B4"] = 'FT2-MP3'
    planilha["D4"] = id   
    planilha["G4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close() 

for filename in os.listdir():
    if filename.startswith('FBR-EN-4-450.xlsx'):
        os.remove('FBR-EN-4-450.xlsx')

for filename in os.listdir():
    if filename.endswith('.xlsx'):
        os.rename(filename, os.path.join('FBR-EN-4-450', filename))
       
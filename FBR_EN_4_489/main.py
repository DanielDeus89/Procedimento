import openpyxl
import os
from openpyxl.utils import get_column_letter
import datetime
import shutil


directory = ''

if os.path.exists("FBR-EN-4-489"):
    shutil.rmtree('FBR-EN-4-489')

if not os.path.exists("FBR-EN-4-489"):
    os.mkdir('FBR-EN-4-489')

print("Estou trabalhando na criação dos arquivos agora. Por favor, aguarde um pouco enquanto eu os crio. Isso pode levar alguns minutos. Obrigado pela sua paciência!")

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".xlsx" e se o nome começa com "SmarX"
    if filename.endswith("xlsx") and filename.startswith("Smar") or filename.startswith("Adda"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # remove o arquivo
        os.remove(current_file)

# percorre todos os arquivos no diretório
for filename in os.listdir():
    # verifica se a extensão do arquivo é ".txt"
    if filename.endswith("xlsx"):
        # constrói o caminho completo do arquivo atual
        current_file = os.path.join(directory, filename)
        # constrói o novo caminho completo do arquivo
        new_file = os.path.join(directory, 'FBR-EN-4-489.xlsx')
        # renomeia o arquivo
        os.rename(current_file, new_file)        


meses = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
mes_atual = datetime.date.today().month

RUNIN = ['RUNIN']
BURN_IN = ['AddaX8754A0140','AddaX8754A0145']

arquivo = openpyxl.load_workbook("FBR-EN-4-489.xlsx")
planilha = arquivo.active
for id in RUNIN:                     
    planilha["C4"] = 'RUNIN' 
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()


arquivo = openpyxl.load_workbook("FBR-EN-4-489.xlsx")
planilha = arquivo.active
for id in BURN_IN:                     
    planilha["C4"] = 'BURN IN' 
    planilha["F4"] = id   
    planilha["AA4"] = meses[mes_atual - 1]  
    arquivo.save(id + ".xlsx")    
arquivo.close()

for filename in os.listdir():
    if filename.startswith('FBR-EN-4-489.xlsx'):
        os.remove('FBR-EN-4-489.xlsx')

for filename in os.listdir():
    if filename.endswith('.xlsx'):
        os.rename(filename, os.path.join('FBR-EN-4-489', filename))
       